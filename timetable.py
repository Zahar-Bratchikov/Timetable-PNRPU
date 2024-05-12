from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
from datetime import datetime
from PyQt5 import QtCore, QtWidgets


class Timetable:
    def __init__(self, filename, sheet_name):
        self.wb = load_workbook(filename)
        self.ws = self.wb[sheet_name]
        self.months = {
            'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4,
            'мая': 5, 'июня': 6, 'июля': 7, 'августа': 8,
            'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
        }

    def extract_date_and_week(self, message):
        parts = message.split()
        day = int(parts[3])
        month = self.months[parts[4].lower()]
        week = int(parts[-2])
        start_date = datetime(datetime.now().year, month, day)
        return start_date, week

    def weeks_since(self, start_date, current_week):
        today = datetime.now()
        delta = today - start_date
        weeks = delta.days // 7 + current_week
        return weeks

    def are_cells_merged(self, cell1_coord, cell2_coord):
        min_row1, min_col1, max_row1, max_col1 = range_boundaries(cell1_coord)
        min_row2, min_col2, max_row2, max_col2 = range_boundaries(cell2_coord)

        for merged_range in self.ws.merged_cells.ranges:
            merged_min_row, merged_min_col, merged_max_row, merged_max_col = merged_range.bounds
            if (min_row1 >= merged_min_row and max_row1 <= merged_max_row and
                    min_col1 >= merged_min_col and max_col1 <= merged_max_col and
                    min_row2 >= merged_min_row and max_row2 <= merged_max_row and
                    min_col2 >= merged_min_col and max_col2 <= merged_max_col):
                return True
        return False

    def get_timetable(self):
        start_date, current_week = self.extract_date_and_week(self.ws['a2'].value)
        now = datetime.now()
        tt_arr = []
        for k in range(0, 2):
            week = k
            for i in range(4, 76):
                if (self.are_cells_merged('c' + str(i - 1), 'c' + str(i)) and name.value is not None):
                    continue
                name = self.ws['c' + str(i)]
                if (i % 2 == week and name.value is not None):
                    tt_arr.append(name.value)
                elif (self.are_cells_merged('c' + str(i), 'c' + str(i + 1)) and name.value is not None):
                    tt_arr.append(name.value)
                elif (i % 2 == week):
                    tt_arr.append("---")
            for i in range(0, 6):
                tt_arr.append("---")


        return tt_arr, now.weekday() + ((self.weeks_since(start_date, current_week) + 1) % 2) * 7, self.ws['c3'].value, self.ws['a1'].value

    def fill_table_data(self, timetable):
        self.current_day = timetable.get_timetable()[2]  # Изменил индекс для получения текущего дня
        self.current_week = timetable.get_timetable()[1]  # Изменил индекс для получения текущей недели

        if self.current_week == 0:
            self.tableWidget.setHorizontalHeaderLabels(["Первая неделя"])
        else:
            self.tableWidget.setHorizontalHeaderLabels(["Вторая неделя"])

        # Получаем данные из расписания
        self.timetable_arr, week, current_day = timetable.get_timetable()

        # Заполняем вертикальные заголовки
        self.tableWidget.setVerticalHeaderLabels(["8:00", "9:40", "11:30", "13:20", "15:00", "16:40"])

        # Обновляем таблицу данными расписания
        for row in range(6):
            for col in range(1):  # Изменил на range(1)
                index = row + col * 6
                text = self.timetable_arr[index]
                item = QtWidgets.QTableWidgetItem(text)
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.tableWidget.setItem(row, col, item)

        # Обновляем метку дня
        days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.lineEdit.setText(days_of_week[current_day])

        # Изменяем высоту строк таблицы
        table_height = self.tableWidget.height()
        row_count = self.tableWidget.rowCount()
        row_height = int(table_height / row_count - 5)
        for row in range(row_count):
            self.tableWidget.setRowHeight(row, row_height)
