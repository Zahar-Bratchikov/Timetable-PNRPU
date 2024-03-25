from openpyxl import *
from openpyxl.utils import *


def are_cells_merged(worksheet, cell1_coord, cell2_coord):
    min_row1, min_col1, max_row1, max_col1 = range_boundaries(cell1_coord)
    min_row2, min_col2, max_row2, max_col2 = range_boundaries(cell2_coord)

    for merged_range in worksheet.merged_cells.ranges:
        merged_min_row, merged_min_col, merged_max_row, merged_max_col = merged_range.bounds
        if (min_row1 >= merged_min_row and max_row1 <= merged_max_row and
                min_col1 >= merged_min_col and max_col1 <= merged_max_col and
                min_row2 >= merged_min_row and max_row2 <= merged_max_row and
                min_col2 >= merged_min_col and max_col2 <= merged_max_col):
            return True
    return False


wb = load_workbook('test.xlsx')
ws = wb["Лист1"]
test = ws['c4']
week = 0
for i in range(4, 75):
    if (i == 4):
        print("Понедельник")
    elif (i == 16):
        print("\nВторник")
    elif (i == 28):
        print("\nСреда")
    elif (i == 40):
        print("\nЧетверг")
    elif (i == 52):
        print("\nПятница")
    elif (i == 64):
        print("\nСуббота")
    if (i % 2 == 0):
        time = ws['b' + str(i)]
        print(time.value)
    name = ws['c' + str(i)]
    if (i % 2 == week):
        if (name.value != None):
            print(name.value)
    elif are_cells_merged(ws, 'c' + str(i), 'c' + str(i + 1)):
        if (name.value != None):
            print(name.value)
